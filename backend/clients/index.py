"""
API базы клиентов FABRICA. v2
GET                  — все клиенты
GET ?id=N            — один клиент + его заказы
POST { type:'upsert_from_order', ...data }  — добавить/обновить из заказа
POST { type:'client', ...data }             — добавить вручную
POST { type:'excel', file: base64 }         — импорт из Excel
PUT  { id, ...fields }                      — обновить клиента
DELETE ?id=N                                — удалить
"""
import os, json, base64
import psycopg2
from psycopg2.extras import RealDictCursor
from decimal import Decimal

CLIENT_FIELDS = ('full_name','phone','email','city','inn','delivery_days','delivery_time','payment_method','delivery_address','delivery_type')

def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'])

def cors():
    return {
        'Access-Control-Allow-Origin': '*',
        'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
        'Access-Control-Allow-Headers': 'Content-Type',
        'Access-Control-Max-Age': '86400',
    }

def to_float(v):
    if v is None: return 0.0
    return float(Decimal(str(v)))

def row_to_client(r):
    return {
        'id': r['id'],
        'full_name': r['full_name'] or '',
        'phone': r['phone'] or '',
        'email': r.get('email') or '',
        'city': r.get('city') or '',
        'inn': r.get('inn') or '',
        'delivery_days': r.get('delivery_days') or '',
        'delivery_time': r.get('delivery_time') or '',
        'payment_method': r.get('payment_method') or '',
        'delivery_address': r.get('delivery_address') or '',
        'delivery_type': r.get('delivery_type') or '',
        'order_count': int(r.get('order_count') or 0),
        'total_sum': int(r.get('total_sum') or 0),
        'created_at': r['created_at'].isoformat() if r.get('created_at') else '',
    }

def handler(event: dict, context) -> dict:
    method = event.get('httpMethod', 'GET')
    if method == 'OPTIONS':
        return {'statusCode': 200, 'headers': cors(), 'body': ''}

    conn = get_conn()
    conn.autocommit = True
    params = event.get('queryStringParameters') or {}

    try:
        if method == 'GET':
            client_id = params.get('id')
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                if client_id:
                    cur.execute("SELECT * FROM clients WHERE id = %s", (int(client_id),))
                    client = cur.fetchone()
                    if not client:
                        return {'statusCode': 404, 'headers': cors(), 'body': json.dumps({'error': 'not found'})}
                    cur.execute(
                        "SELECT id, order_number, stage, city, customer_name, total, discount, "
                        "items, created_at, due_date, notes "
                        "FROM orders WHERE phone = %s AND is_trashed = FALSE ORDER BY created_at DESC",
                        (client['phone'],)
                    )
                    orders = []
                    for r in cur.fetchall():
                        items = r['items'] if isinstance(r['items'], list) else []
                        orders.append({
                            'id': r['id'], 'order_number': r['order_number'],
                            'stage': r['stage'], 'city': r['city'] or '',
                            'customer_name': r['customer_name'] or '',
                            'total': r['total'], 'discount': r['discount'] or 0,
                            'items': items,
                            'created_at': r['created_at'].isoformat() if r['created_at'] else '',
                            'due_date': r['due_date'].isoformat() if r['due_date'] else '',
                            'notes': r['notes'] or '',
                        })
                    c = row_to_client({**dict(client), 'order_count': len(orders), 'total_sum': sum(o['total'] for o in orders)})
                    return {'statusCode': 200, 'headers': cors(), 'body': json.dumps({'client': c, 'orders': orders})}
                else:
                    cur.execute("""
                        SELECT c.*, COUNT(o.id) as order_count,
                               COALESCE(SUM(o.total), 0) as total_sum
                        FROM clients c
                        LEFT JOIN orders o ON o.phone = c.phone AND o.is_trashed = FALSE
                        GROUP BY c.id ORDER BY total_sum DESC
                    """)
                    clients = [row_to_client(r) for r in cur.fetchall()]
                    return {'statusCode': 200, 'headers': cors(), 'body': json.dumps({'clients': clients})}

        body = json.loads(event.get('body') or '{}')

        if method == 'POST':
            b_type = body.get('type', 'client')

            if b_type == 'upsert_from_order':
                phone = (body.get('phone') or '').strip()
                if not phone:
                    return {'statusCode': 200, 'headers': cors(), 'body': json.dumps({'skipped': True})}
                full_name = body.get('full_name', '').strip()
                email     = (body.get('email') or '').strip()
                city      = (body.get('city') or '').strip()
                inn       = (body.get('inn') or '').strip()
                delivery_days    = (body.get('delivery_days') or '').strip()
                delivery_time    = (body.get('delivery_time') or '').strip()
                payment_method   = (body.get('payment_method') or '').strip()
                delivery_address = (body.get('delivery_address') or '').strip()
                delivery_type    = (body.get('delivery_type') or '').strip()

                with conn.cursor() as cur:
                    cur.execute("SELECT id FROM clients WHERE phone = %s", (phone,))
                    existing = cur.fetchone()
                    if existing:
                        updates, vals = [], []
                        for f, v in [('full_name',full_name),('email',email),('city',city),
                                     ('inn',inn),('delivery_days',delivery_days),
                                     ('delivery_time',delivery_time),('payment_method',payment_method),
                                     ('delivery_address',delivery_address),('delivery_type',delivery_type)]:
                            if v:
                                updates.append(f'{f} = %s'); vals.append(v)
                        if updates:
                            updates.append('updated_at = NOW()')
                            vals.append(existing[0])
                            cur.execute(f"UPDATE clients SET {', '.join(updates)} WHERE id = %s", vals)
                        return {'statusCode': 200, 'headers': cors(), 'body': json.dumps({'id': existing[0], 'new': False})}
                    else:
                        cur.execute(
                            "INSERT INTO clients (full_name,phone,email,city,inn,delivery_days,delivery_time,payment_method,delivery_address,delivery_type) "
                            "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                            (full_name or phone, phone, email, city, inn, delivery_days, delivery_time, payment_method, delivery_address, delivery_type)
                        )
                        new_id = cur.fetchone()[0]
                        return {'statusCode': 200, 'headers': cors(), 'body': json.dumps({'id': new_id, 'new': True})}

            if b_type == 'excel':
                # Импорт из Excel (base64)
                try:
                    import io
                    try:
                        import openpyxl
                    except ImportError:
                        return {'statusCode': 500, 'headers': cors(), 'body': json.dumps({'error': 'openpyxl not installed'})}
                    file_b64 = body.get('file', '')
                    file_bytes = base64.b64decode(file_b64)
                    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
                    ws = wb.active
                    headers = [str(c.value or '').strip().lower() for c in ws[1]]
                    created = updated = 0
                    with conn.cursor() as cur:
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            r = dict(zip(headers, row))
                            phone = str(r.get('телефон') or r.get('phone') or '').strip()
                            if not phone: continue
                            full_name = str(r.get('имя') or r.get('full_name') or r.get('название') or phone).strip()
                            email     = str(r.get('email') or r.get('почта') or '').strip()
                            city      = str(r.get('город') or r.get('city') or '').strip()
                            inn       = str(r.get('инн') or r.get('inn') or '').strip()
                            cur.execute("SELECT id FROM clients WHERE phone = %s", (phone,))
                            ex = cur.fetchone()
                            if ex:
                                cur.execute("UPDATE clients SET full_name=%s,email=%s,city=%s,inn=%s,updated_at=NOW() WHERE id=%s",
                                            (full_name, email, city, inn, ex[0]))
                                updated += 1
                            else:
                                cur.execute("INSERT INTO clients (full_name,phone,email,city,inn) VALUES (%s,%s,%s,%s,%s)",
                                            (full_name, phone, email, city, inn))
                                created += 1
                    return {'statusCode': 200, 'headers': cors(), 'body': json.dumps({'created': created, 'updated': updated})}
                except Exception as e:
                    return {'statusCode': 500, 'headers': cors(), 'body': json.dumps({'error': str(e)})}

            if b_type == 'client':
                phone = (body.get('phone') or '').strip()
                if not phone:
                    return {'statusCode': 400, 'headers': cors(), 'body': json.dumps({'error': 'phone required'})}
                cols, vals = ['full_name','phone'], [body.get('full_name', phone).strip(), phone]
                for f in ('email','city','inn','delivery_days','delivery_time','payment_method','delivery_address','delivery_type'):
                    if f in body and body[f] is not None:
                        cols.append(f); vals.append(str(body[f]).strip())
                placeholders = ','.join(['%s']*len(cols))
                update_set = ','.join(f'{c}=EXCLUDED.{c}' for c in cols if c != 'phone')
                with conn.cursor() as cur:
                    cur.execute(
                        f"INSERT INTO clients ({','.join(cols)}) VALUES ({placeholders}) "
                        f"ON CONFLICT (phone) DO UPDATE SET {update_set}, updated_at=NOW() RETURNING id",
                        vals
                    )
                    new_id = cur.fetchone()[0]
                return {'statusCode': 200, 'headers': cors(), 'body': json.dumps({'id': new_id})}

            # sync_from_orders — заполнить клиентов из всех заказов
            if b_type == 'sync_from_orders':
                with conn.cursor(cursor_factory=RealDictCursor) as cur:
                    cur.execute("SELECT customer_name, phone, customer_email, city, form, payment_method, delivery_address, delivery_type FROM orders WHERE phone IS NOT NULL AND phone <> '' AND is_trashed = FALSE")
                    rows = cur.fetchall()
                synced = 0
                for r in rows:
                    phone = (r['phone'] or '').strip()
                    if not phone: continue
                    form = r['form'] or {}
                    with conn.cursor() as cur:
                        cur.execute("SELECT id FROM clients WHERE phone = %s", (phone,))
                        ex = cur.fetchone()
                        full_name = (r['customer_name'] or '').strip() or phone
                        email     = (r['customer_email'] or form.get('email') or '').strip()
                        city      = (r['city'] or '').strip()
                        inn       = (form.get('inn') or '').strip()
                        d_days    = (form.get('delivery_days') or '').strip()
                        d_time    = (form.get('delivery_time') or '').strip()
                        pmeth     = (r['payment_method'] or '').strip()
                        daddr     = (r['delivery_address'] or '').strip()
                        dtype     = (r['delivery_type'] or '').strip()
                        if ex:
                            cur.execute(
                                "UPDATE clients SET full_name=%s,email=%s,city=%s,inn=%s,delivery_days=%s,delivery_time=%s,payment_method=%s,delivery_address=%s,delivery_type=%s,updated_at=NOW() WHERE id=%s",
                                (full_name, email, city, inn, d_days, d_time, pmeth, daddr, dtype, ex[0])
                            )
                        else:
                            cur.execute(
                                "INSERT INTO clients (full_name,phone,email,city,inn,delivery_days,delivery_time,payment_method,delivery_address,delivery_type) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                (full_name, phone, email, city, inn, d_days, d_time, pmeth, daddr, dtype)
                            )
                        synced += 1
                return {'statusCode': 200, 'headers': cors(), 'body': json.dumps({'synced': synced})}

        if method == 'PUT':
            client_id = int(body.get('id'))
            fields, vals = [], []
            for col in CLIENT_FIELDS:
                if col in body:
                    fields.append(f'{col} = %s'); vals.append(body[col])
            if fields:
                fields.append('updated_at = NOW()')
                vals.append(client_id)
                with conn.cursor() as cur:
                    cur.execute(f"UPDATE clients SET {', '.join(fields)} WHERE id = %s", vals)
            return {'statusCode': 200, 'headers': cors(), 'body': json.dumps({'ok': True})}

        if method == 'DELETE':
            client_id = int(params.get('id'))
            with conn.cursor() as cur:
                cur.execute("DELETE FROM clients WHERE id = %s", (client_id,))
            return {'statusCode': 200, 'headers': cors(), 'body': json.dumps({'ok': True})}

        return {'statusCode': 405, 'headers': cors(), 'body': ''}

    except Exception as e:
        return {'statusCode': 500, 'headers': cors(), 'body': json.dumps({'error': str(e)})}
    finally:
        conn.close()
