"""Выгрузка всех товаров в Excel файл (base64)."""
import json
import os
import base64
import io
import psycopg2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

CORS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type',
}

COLUMNS = [
    ('артикул',          'sku'),
    ('название',         'name'),
    ('описание',         'description'),
    ('форма',            'shape'),
    ('размер',           'size'),
    ('цвет',             'color'),
    ('цена',             'price'),
    ('цена по акции',    'sale_price'),
    ('фото',             'image_url'),
    ('группа',           'group_id'),
    ('группировать по',  'group_by'),
    ('разделить по',     'split_by'),
    ('метки',            'labels'),
    ('приоритет',        'priority'),
    ('вид плетения',     'weave_type'),
    ('кол-во ручек',     'handles_count'),
]

def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'], options=f"-c search_path={os.environ['MAIN_DB_SCHEMA']}")

def handler(event: dict, context) -> dict:
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS, 'body': ''}

    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT sku, name, description, shape, size, color, price, sale_price,
               image_url, group_id, group_by, split_by, labels, priority, weave_type, handles_count
        FROM products ORDER BY priority NULLS LAST, id
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Товары'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='5A3E28')

    for col_idx, (col_name, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode()

    return {
        'statusCode': 200,
        'headers': {**CORS, 'Content-Type': 'application/json'},
        'body': json.dumps({'file': b64, 'filename': 'products.xlsx'})
    }
