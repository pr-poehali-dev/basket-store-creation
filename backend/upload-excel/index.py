"""Загрузка товаров из Excel. При наличии артикула (sku) — обновляет существующий товар, иначе создаёт новый."""
import json
import os
import base64
import io
import re
import psycopg2
import openpyxl

CORS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'POST, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type',
}

VALID_SHAPES = ['Круглые', 'Овальные', 'Прямоугольные', 'Сердечки']

def get_size_category(size_str: str) -> str:
    if not size_str:
        return ''
    nums = re.findall(r'\d+', size_str.replace(',', '.'))
    if not nums:
        return ''
    first = int(nums[0])
    if first < 30:
        return 'до 30 см'
    elif first <= 50:
        return '30-50 см'
    else:
        return '50-120 см'

def get_conn():
    return psycopg2.connect(os.environ['DATABASE_URL'], options=f"-c search_path={os.environ['MAIN_DB_SCHEMA']}")

def handler(event: dict, context) -> dict:
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS, 'body': ''}

    body = json.loads(event.get('body') or '{}')
    file_b64 = body.get('file')
    mode = body.get('mode', 'append')

    if not file_b64:
        return {'statusCode': 400, 'headers': CORS, 'body': json.dumps({'error': 'Файл не передан'})}

    file_bytes = base64.b64decode(file_b64)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def col(row, *names):
        for name in names:
            if name in headers:
                v = row[headers.index(name)].value
                return str(v).strip() if v is not None else ''
        return ''

    rows_data = []
    for row in ws.iter_rows(min_row=2):
        if not any(c.value for c in row):
            continue
        name = col(row, 'название', 'name')
        if not name:
            continue
        shape = col(row, 'форма', 'shape') or 'Круглые'
        if shape not in VALID_SHAPES:
            shape = 'Круглые'
        size = col(row, 'размер', 'size') or ''
        size_category = get_size_category(size)
        try:
            price = int(float(col(row, 'цена', 'price') or 0))
        except:
            price = 0
        sale_price_raw = col(row, 'цена по акции', 'sale_price', 'акция')
        try:
            sale_price = int(float(sale_price_raw)) if sale_price_raw else None
        except:
            sale_price = None
        priority_raw = col(row, 'приоритет', 'priority')
        try:
            priority = int(float(priority_raw)) if priority_raw else None
        except:
            priority = None

        rows_data.append({
            'sku': col(row, 'артикул', 'sku') or None,
            'name': name,
            'description': col(row, 'описание', 'description'),
            'shape': shape,
            'size': size,
            'size_category': size_category,
            'color': col(row, 'цвет', 'color'),
            'price': price,
            'sale_price': sale_price,
            'image_url': col(row, 'фото', 'image_url'),
            'group_id': col(row, 'группа', 'group_id') or None,
            'group_by': col(row, 'группировать по', 'group_by') or None,
            'split_by': col(row, 'разделить по', 'split_by') or None,
            'labels': col(row, 'метки', 'labels') or None,
            'priority': priority,
            'weave_type': col(row, 'вид плетения', 'weave_type') or None,
            'handles_count': col(row, 'кол-во ручек', 'handles_count') or None,
        })

    conn = get_conn()
    cur = conn.cursor()

    if mode == 'replace':
        cur.execute("DELETE FROM products")

    imported = 0
    for p in rows_data:
        if p['sku'] and mode != 'replace':
            # Upsert по артикулу
            cur.execute("SELECT id FROM products WHERE sku=%s", (p['sku'],))
            existing = cur.fetchone()
            if existing:
                cur.execute(
                    """UPDATE products SET name=%s, description=%s, shape=%s, size=%s, size_category=%s,
                       color=%s, price=%s, sale_price=%s, image_url=%s, group_id=%s, group_by=%s,
                       split_by=%s, labels=%s, priority=%s, weave_type=%s, handles_count=%s,
                       updated_at=NOW() WHERE id=%s""",
                    (p['name'], p['description'], p['shape'], p['size'], p['size_category'],
                     p['color'], p['price'], p['sale_price'], p['image_url'],
                     p['group_id'], p['group_by'], p['split_by'],
                     p['labels'], p['priority'], p['weave_type'], p['handles_count'], existing[0])
                )
                imported += 1
                continue
        cur.execute(
            """INSERT INTO products (sku, name, description, shape, size, size_category, color, price, sale_price,
               image_url, group_id, group_by, split_by, labels, priority, weave_type, handles_count)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (p['sku'], p['name'], p['description'], p['shape'], p['size'], p['size_category'],
             p['color'], p['price'], p['sale_price'], p['image_url'],
             p['group_id'], p['group_by'], p['split_by'],
             p['labels'], p['priority'], p['weave_type'], p['handles_count'])
        )
        imported += 1

    conn.commit()
    cur.close()
    conn.close()

    return {'statusCode': 200, 'headers': CORS, 'body': json.dumps({'imported': imported})}
